#!/usr/bin/env python3
import sys, json, io, os, tempfile, requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

BLACK = 'FF000000'

def _s(style=None):
    return Side(border_style=style, color=BLACK) if style else Side(border_style=None)

MED = _s('medium')
NO  = _s(None)

def set_outer_border(ws, min_row, max_row, min_col, max_col):
    """Black medium border on outer perimeter of a cell range (no merge needed)."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            left   = MED if c == min_col  else NO
            right  = MED if c == max_col  else NO
            top    = MED if r == min_row  else NO
            bottom = MED if r == max_row  else NO
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def download_image(url):
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        return r.content
    except:
        return None

def crop_fill(pil_img, w, h):
    """Center-crop then resize to exactly w x h."""
    ow, oh = pil_img.size
    if ow / oh > w / h:
        nw = int(oh * w / h)
        pil_img = pil_img.crop(((ow - nw)//2, 0, (ow + nw)//2, oh))
    else:
        nh = int(ow * h / w)
        pil_img = pil_img.crop((0, (oh - nh)//2, ow, (oh + nh)//2))
    return pil_img.resize((w, h), PILImage.LANCZOS)

def make_laporan(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    ws.column_dimensions['A'].width = 0.5
    for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N']:
        ws.column_dimensions[col].width = 9.7109375

    label_rows = {5,6,18,19,31,32,44,45}
    photo_rows = set(range(7,18))|set(range(20,31))|set(range(33,44))|set(range(46,57))
    for r in range(1, 70):
        ws.row_dimensions[r].height = 16 if r in label_rows else (20 if r in photo_rows else 15)

    # Header
    ws.merge_cells('B1:C4')
    ws['B1'].value     = 'Logo'
    ws['B1'].font      = Font(name='Calibri', size=11)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    set_outer_border(ws, 1, 4, 2, 3)   # B1:C4

    ws.merge_cells('D1:K4')
    ws['D1'].value     = data.get('judul', 'LAPORAN INSPEKSI PEKERJAAN')
    ws['D1'].font      = Font(name='Calibri', bold=True, size=16)
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    set_outer_border(ws, 1, 4, 4, 11)  # D1:K4

    ws.merge_cells('L1:M4')
    ws['L1'].alignment = Alignment(horizontal='center', vertical='center')
    set_outer_border(ws, 1, 4, 12, 13) # L1:M4

    # Approximate pixel size per cell
    # col: 9.71 char * 7px = ~68px, row: 20pt * 96/72 = ~26.7px
    COL_PX = 68
    ROW_PX = 24   # lebih konservatif agar tidak overflow
    BOX_COLS = 4
    BOX_ROWS = 11
    BOX_W = COL_PX * BOX_COLS  # 272px
    BOX_H = ROW_PX * BOX_ROWS  # 264px

    sections = [
        {'dr': (5, 6),   'pr': (7, 17)},
        {'dr': (18, 19), 'pr': (20, 30)},
        {'dr': (31, 32), 'pr': (33, 43)},
        {'dr': (44, 45), 'pr': (46, 56)},
    ]

    # Tiap sisi: outer_min/max = range dengan border, photo_min/max = range yang di-merge untuk foto
    sides = [
        # Kiri: outer B:G (2-7), foto merge C:F (3-6), anchor di C
        {'outer_min': 2,  'outer_max': 7,  'photo_min': 3,  'photo_max': 6,  'anchor': 'C'},
        # Kanan: outer H:M (8-13), foto merge I:L (9-12), anchor di I
        {'outer_min': 8,  'outer_max': 13, 'photo_min': 9,  'photo_max': 12, 'anchor': 'I'},
    ]

    foto_paths  = data.get('foto_paths',    [])
    foto_descs  = data.get('foto_deskripsis', [])
    temp_files  = []
    foto_idx    = 0

    for sec in sections:
        dr1, dr2 = sec['dr']
        pr1, pr2 = sec['pr']

        for sd in sides:
            outer_min = sd['outer_min']
            outer_max = sd['outer_max']
            photo_min = sd['photo_min']
            photo_max = sd['photo_max']
            anchor    = sd['anchor']

            desc_text = foto_descs[foto_idx] if foto_idx < len(foto_descs) else ''

            # Outer box border (B:G atau H:M) tanpa merge
            set_outer_border(ws, dr1, pr2, outer_min, outer_max)

            # Description row – merge seluruh lebar outer box
            outer_anchor_letter = chr(ord('A') + outer_min - 1)  # kolom B atau H
            outer_end_letter    = chr(ord('A') + outer_max - 1)  # kolom G atau M
            desc_range = f'{outer_anchor_letter}{dr1}:{outer_end_letter}{dr2}'
            ws.merge_cells(desc_range)
            dc = ws[f'{outer_anchor_letter}{dr1}']
            dc.value     = f'Deskripsi :  {desc_text}'
            dc.font      = Font(name='Calibri', size=12)
            dc.alignment = Alignment(horizontal='left', vertical='center')

            # Foto box – merge C:F atau I:L saja (4 kolom tengah)
            photo_range = f'{anchor}{pr1}:{chr(ord(anchor)+3)}{pr2}'
            ws.merge_cells(photo_range)
            pc = ws[f'{anchor}{pr1}']
            pc.alignment = Alignment(horizontal='center', vertical='center')

            # Insert image
            if foto_idx < len(foto_paths) and foto_paths[foto_idx]:
                img_bytes = download_image(foto_paths[foto_idx])
                if img_bytes:
                    try:
                        pil = PILImage.open(io.BytesIO(img_bytes)).convert('RGB')
                        # thumbnail: resize proporsional, tidak melebihi BOX
                        pil.thumbnail((BOX_W, BOX_H), PILImage.LANCZOS)
                        img_w, img_h = pil.size

                        tmp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                        pil.save(tmp.name, 'JPEG', quality=85)
                        tmp.close()
                        temp_files.append(tmp.name)

                        xl_img        = XLImage(tmp.name)
                        xl_img.width  = img_w
                        xl_img.height = img_h
                        xl_img.anchor = f'{anchor}{pr1}'
                        ws.add_image(xl_img)
                    except Exception as e:
                        sys.stderr.write(f'Foto error slot {foto_idx}: {e}\n')
                        pc.value = '[Foto tidak tersedia]'
                else:
                    pc.value = '[Foto tidak dapat diunduh]'

            foto_idx += 1

    # Logo kanan atas
    logo_url = data.get('logo_url', '')
    if logo_url:
        logo_bytes = download_image(logo_url)
        if logo_bytes:
            try:
                pil = PILImage.open(io.BytesIO(logo_bytes)).convert('RGBA')
                pil.thumbnail((173, 173), PILImage.LANCZOS)
                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                pil.save(tmp.name, 'PNG')
                tmp.close()
                temp_files.append(tmp.name)
                xl_logo = XLImage(tmp.name)
                xl_logo.width  = 173
                xl_logo.height = 173
                xl_logo.anchor = 'L1'
                ws.add_image(xl_logo)
            except: pass

    # Footer
    tanggal = data.get('tanggal', 'dd-mm-yy')
    ws.merge_cells('K57:K58')
    ws['K57'].value     = 'Jakarta,'
    ws['K57'].font      = Font(name='Calibri', bold=True, size=11)
    ws['K57'].alignment = Alignment(horizontal='right', vertical='center')
    ws['K57'].border    = Border(top=MED)
    ws.merge_cells('L57:M58')
    ws['L57'].value     = tanggal
    ws['L57'].font      = Font(name='Calibri', bold=True, size=11)
    ws['L57'].alignment = Alignment(horizontal='left', vertical='center')
    ws['L57'].border    = Border(top=MED)

    for ref, txt in [('B59','Diajukan Oleh,'),('F59','Diperiksa Oleh,'),('K59','Diperiksa Oleh,')]:
        ws[ref].value = txt
        ws[ref].font  = Font(name='Calibri', size=11)

    for rng in ['B63:D63', 'F63:H63', 'K63:M63']:
        ws.merge_cells(rng)
        ws[rng.split(':')[0]].border = Border(bottom=MED)

    for ref, key, default in [
        ('B65','diajukan_oleh','CV. Garuda Jaya'),
        ('F65','diperiksa_oleh_1','CV. Titian Mahakarya'),
        ('K65','diperiksa_oleh_2','PT.'),
    ]:
        ws[ref].value = data.get(key, default)
        ws[ref].font  = Font(name='Calibri', bold=True, size=11)

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = 'A1:N65'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    for f in temp_files:
        try: os.unlink(f)
        except: pass

    return out.read()


if __name__ == '__main__':
    try:
        raw = sys.stdin.read() if len(sys.argv) < 2 else sys.argv[1]
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.stderr.write(f"JSON error: {e}\n")
        sys.exit(1)

    sys.stdout.buffer.write(make_laporan(data))
